import os

# Removed gevent monkey patching - using sync workers for better SQLAlchemy compatibility
if os.getenv("FLASK_ENV") == "development":
    print("INFO: Using sync workers (no gevent monkey patching)")

# Import psycopg2 to register PostgreSQL dialect with SQLAlchemy
try:
    import psycopg2
    # Explicitly register the PostgreSQL dialect for SQLAlchemy
    from sqlalchemy.dialects import postgresql

    # Try to manually register the dialect if needed
    try:
        from sqlalchemy.dialects import registry
        if hasattr(registry, '_load'):
            # SQLAlchemy 1.4 style
            registry._load('postgresql', 'sqlalchemy.dialects.postgresql')
        print("INFO: psycopg2 and PostgreSQL dialect imported successfully")
    except Exception as reg_e:
        print(f"INFO: psycopg2 imported, dialect registration skipped: {reg_e}")

except ImportError:
    print("ERROR: psycopg2 not available for PostgreSQL support")

import time
import uuid
import pytz
import re
import logging
from datetime import datetime, timedelta
from urllib.parse import urlparse

from dotenv import load_dotenv
from flask import Flask, current_app, render_template, request, redirect, url_for, flash, session, jsonify, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func, or_
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from flask_migrate import Migrate
from read_logs import bp as read_logs_bp
from flask_login import LoginManager
from models import User, Update, ReadLog, SOPSummary, LessonLearned, ActivityLog, ArchivedUpdate, ArchivedSOPSummary, ArchivedLessonLearned
from extensions import db, socketio
from database import db_session
from role_decorators import admin_required, editor_required, writer_required, delete_required, export_required, get_user_role_info
from timezone_utils import UTC, IST, now_utc, to_utc, to_ist, format_ist, ensure_timezone, is_within_hours, get_hours_ago
from io import BytesIO
import sys
import subprocess
from pathlib import Path

# Performance monitoring
def performance_logger(f):
    """Decorator to log response times for performance monitoring"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        start_time = time.time()
        try:
            result = f(*args, **kwargs)
            end_time = time.time()
            duration = end_time - start_time

            # Log slow requests (>1 second)
            if duration > 1.0:
                logger.warning(".2f"
                              f"endpoint={request.endpoint} "
                              f"method={request.method} "
                              f"path={request.path} "
                              f"user_agent={request.headers.get('User-Agent', 'Unknown')[:50]}")

            # Log all requests for performance analysis
            logger.info(".3f"
                       f"endpoint={request.endpoint} "
                       f"method={request.method} "
                       f"status={getattr(result, 'status_code', 'N/A')} "
                       f"size={getattr(result, 'content_length', 'N/A')}")

            return result
        except Exception as e:
            end_time = time.time()
            duration = end_time - start_time
            logger.error(".2f"
                        f"endpoint={request.endpoint} "
                        f"method={request.method} "
                        f"error={str(e)}")
            raise
    return decorated_function

# Lightweight CSV export available for light version
EXCEL_EXPORT_AVAILABLE = True  # CSV export implemented for light version

# Load .env
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Logging configuration removed for Vercel deployment

migrate = Migrate()
login_manager = LoginManager()
login_manager.login_view = 'login'

# Process start timestamp for basic metrics
APP_START = time.time()

# Simple in-memory cache for performance optimization
_cache = {}
CACHE_TIMEOUT = 300  # 5 minutes

def get_cached_user_role(user_id):
    """Get cached user role information"""
    cache_key = f"user_role_{user_id}"
    if cache_key in _cache:
        cached_time, role_info = _cache[cache_key]
        if time.time() - cached_time < CACHE_TIMEOUT:
            return role_info
        else:
            del _cache[cache_key]

    user = User.query.get(user_id)
    if user:
        role_info = get_user_role_info(user)
        _cache[cache_key] = (time.time(), role_info)
        return role_info

    return {'is_admin': False, 'is_editor': False, 'is_writer': False, 'is_deleter': False, 'is_exporter': False}

def create_app(config_name=None):
    app = Flask(__name__)

    # Database configuration - ensure PostgreSQL is used consistently
    database_url = os.getenv('DATABASE_URL')

    # For Vercel, DATABASE_URL should be set from environment variables
    if not database_url:
        # For local development without DATABASE_URL, use SQLite as fallback
        if os.getenv("FLASK_ENV") == "development":
            database_url = "sqlite:///loopin_dev.db"
            print("WARNING: No DATABASE_URL set, using SQLite for local development")
        else:
            # For production (Vercel), DATABASE_URL is required
            database_url = "postgresql://localhost/loopin_dev"  # This will fail but shows the requirement

    # For SQLAlchemy 2.0 compatibility, ensure the database URL uses the correct format
    if database_url.startswith('postgres://'):
        # Convert postgres:// to postgresql:// for SQLAlchemy 2.0
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
        print(f"INFO: Converted database URL to: {database_url}")

    # Clean up invalid connection parameters that can cause psycopg2 errors
    from database import clean_database_url
    database_url = clean_database_url(database_url)

    # Set the cleaned URL in config BEFORE any database operations
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev')
    
    # Apply custom configuration
    if config_name:
        app.config.from_object(config_name)
    
    # Initialize extensions
    migrate.init_app(app, db)
    login_manager.init_app(app)
    
    # Initialize Socket.IO with optimized configuration
    socketio_kwargs = {
        'message_queue': None,
        'cors_allowed_origins': '*',
        'ping_timeout': 30000,  # 30 seconds
        'ping_interval': 15000,  # 15 seconds
        'max_http_buffer_size': 500000,
        'async_mode': 'threading',
        'transports': ['websocket', 'polling'],
        'compression': True,
        'compression_threshold': 512,
        'path': '/socket.io',  # Ensure path matches client
        'allow_upgrades': True,
        'cookie': False
    }

    # Initialize Socket.IO blueprint first
    try:
        from api.socketio import bp as socketio_bp, init_socketio
        app.register_blueprint(socketio_bp)
        # Initialize Socket.IO with proper configuration before init_app
        init_socketio(socketio, app)
        if os.getenv("FLASK_ENV") == "development":
            print("SUCCESS: Socket.IO blueprint registered successfully")
    except Exception as e:
        # Socket.IO registration should not break app startup if something is off
        logger.error(f"Socket.IO blueprint registration failed: {e}")
        if os.getenv("FLASK_ENV") == "development":
            print(f"WARNING: Socket.IO blueprint registration failed: {e}")
            import traceback
            print(f"WARNING: Full error: {traceback.format_exc()}")
        pass

    socketio.init_app(app, **socketio_kwargs)
    app.register_blueprint(read_logs_bp)
    # Register new API blueprint (first milestone)
    try:
        from api.updates import updates_bp as api_bp
        app.register_blueprint(api_bp)
    except Exception:
        # Blueprint registration should not break app startup if something is off
        pass
    
    # Register search API blueprint
    try:
        from api.search import bp as search_bp
        app.register_blueprint(search_bp)
    except Exception:
        # Search API registration should not break app startup if something is off
        pass

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    # Security configuration
    app.secret_key = os.getenv("FLASK_SECRET_KEY", "replace-this-with-a-secure-random-string")
    app.config["APP_NAME"] = "LoopIn"

    # CORS configuration for Socket.IO
    app.config["CORS_HEADERS"] = "Content-Type"

    # Secure session configuration - adjust for Vercel
    is_production = os.getenv("FLASK_ENV") == "production"
    if is_production:
        app.config["SESSION_COOKIE_SECURE"] = True  # Only send cookies over HTTPS
    else:
        app.config["SESSION_COOKIE_SECURE"] = False  # Allow HTTP for development
    app.config["SESSION_COOKIE_HTTPONLY"] = True  # Prevent JavaScript access to session cookie
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"  # CSRF protection
    app.config["PERMANENT_SESSION_LIFETIME"] = 3600  # Session timeout in seconds

    # Vercel-specific configurations
    if is_production:
        app.config["SERVER_NAME"] = None  # Disable SERVER_NAME for Vercel
        app.config["PREFERRED_URL_SCHEME"] = "https"  # Force HTTPS on Vercel

    # Database configuration - DATABASE_URL should be set from the first configuration block
    database_url = app.config['SQLALCHEMY_DATABASE_URI']
    if os.getenv("FLASK_ENV") == "development":
        print(f"Using database: {database_url}")

    # Skip database type validation for local development with SQLite fallback
    flask_env = os.getenv("FLASK_ENV")
    database_url = os.getenv('DATABASE_URL')

    # Clean up invalid connection parameters that can cause psycopg2 errors
    if database_url:
        from database import clean_database_url
        database_url = clean_database_url(database_url)

    if not (flask_env == "development" and not database_url):
        from database import validate_database_type
        if not validate_database_type():
            logger.warning("Database type validation failed - using fallback configuration")

    # Additional validation for SQLite detection
    parsed = urlparse(database_url)
    if parsed.scheme in ("sqlite", "sqlite3"):
        logger.warning("WARNING: Using SQLite - this may cause issues in production")

    parsed = urlparse(database_url)
    if parsed.scheme not in ("postgresql", "postgres", "sqlite", "sqlite3"):
        raise RuntimeError(f"Unsupported DB scheme: {parsed.scheme}")

    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    # Configure SSL for PostgreSQL with Vercel-friendly defaults
    if parsed.scheme in ("postgresql", "postgres"):
        ssl_config = {}

        # Set SSL mode with Vercel-friendly defaults
        ssl_mode = os.getenv("PG_SSLMODE", "require")  # Default to 'require' for Vercel
        ssl_config["sslmode"] = ssl_mode
        logger.info(f"PostgreSQL SSL mode: {ssl_mode}")

        # Vercel-specific SSL parameters
        if os.getenv("PG_SSLROOTCERT"):
            ssl_config["sslrootcert"] = os.getenv("PG_SSLROOTCERT")
        if os.getenv("PG_SSLCERT"):
            ssl_config["sslcert"] = os.getenv("PG_SSLCERT")
        if os.getenv("PG_SSLKEY"):
            ssl_config["sslkey"] = os.getenv("PG_SSLKEY")

        # Set connection arguments for Vercel with optimized connection pooling
        app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
            "connect_args": ssl_config,
            # Optimized connection pool settings for Vercel deployment
            "pool_pre_ping": True,
            "pool_recycle": 300,  # Recycle connections every 5 minutes
            "pool_timeout": 20,   # Connection timeout
            "pool_size": 3,       # Base pool size for Vercel
            "max_overflow": 5,    # Maximum overflow
            # Transaction isolation and connection stability
            "isolation_level": "READ_COMMITTED",  # Explicit isolation level
            "echo": False,  # Disable SQL echoing in production
            "pool_reset_on_return": "rollback",  # Reset connections on return
            # Additional stability settings
            "connect_args": {
                **ssl_config,
                "application_name": "loopin-core",  # Identify connections
                "keepalives": 1,  # Enable TCP keepalives
                "keepalives_idle": 30,  # TCP keepalive settings
                "keepalives_interval": 10,
                "keepalives_count": 5,
            }
        }

        logger.info(f"SQLAlchemy engine options configured for Vercel PostgreSQL")
    else:
        # Non-PostgreSQL databases use optimized settings
        app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
            "pool_pre_ping": True,
            "pool_recycle": 300,
            "pool_timeout": 30,
            "pool_size": 10,
            "max_overflow": 20,
            "pool_reset_on_return": "rollback",
        }

    # Ensure PostgreSQL dialect is loaded before initializing the database
    try:
        from sqlalchemy.dialects import postgresql
        print("INFO: PostgreSQL dialect loaded before db.init_app")

        # Try to manually register the dialect if needed
        from sqlalchemy.engine import url
        from sqlalchemy import pool
        print("INFO: Attempting manual dialect registration")

    except ImportError as e:
        print(f"WARNING: Could not load PostgreSQL dialect: {e}")

    # Ensure the database URL is clean before initializing the database
    if 'supa=' in app.config['SQLALCHEMY_DATABASE_URI']:
        from database import clean_database_url
        app.config['SQLALCHEMY_DATABASE_URI'] = clean_database_url(app.config['SQLALCHEMY_DATABASE_URI'])

    # Dispose of any existing engine to ensure it uses the cleaned URL
    try:
        if hasattr(db, '_engine') and db._engine:
            db._engine.dispose()
    except:
        pass

    db.init_app(app)

    with app.app_context():
        if os.getenv("FLASK_ENV") == "development":
            print("Database configured successfully")

        # Verify database connection and table reflection
        db_verification_success = False
        try:
            # Skip database verification in development mode to avoid dialect issues
            if os.getenv("FLASK_ENV") == "development":
                logger.info("Skipping database verification in development mode")
                db_verification_success = True
            else:
                # Comprehensive database health check for production
                from database import health_check
                if not health_check():
                    logger.error("Database health check failed")
                    raise RuntimeError("Database health check failed")

                # Test basic connection with health check
                from database import get_connection_with_retry
                with get_connection_with_retry() as test_conn:
                    test_conn.execute(text("SELECT 1"))
                logger.info("Database connection successful")

            # Test SSL connection if using PostgreSQL (in separate session to avoid transaction abortion)
            ssl_test_result = False
            try:
                if os.getenv("FLASK_ENV") != "development":
                    from database import test_ssl_connection
                    # Create a new healthy connection for SSL testing
                    with get_connection_with_retry() as test_conn:
                        ssl_test_result = test_ssl_connection(test_conn)
                    if ssl_test_result:
                        logger.info("SSL connection test passed")
                    else:
                        logger.warning("SSL connection test failed - this may cause issues")
                else:
                    logger.info("Skipping SSL test in development mode")
            except Exception as ssl_e:
                logger.warning(f"SSL test error (non-critical): {ssl_e}")

            # Check if tables exist (skip in development mode to avoid engine issues)
            if os.getenv("FLASK_ENV") != "development":
                from sqlalchemy import inspect
                inspector = inspect(db.engine)
                existing_tables = inspector.get_table_names()
                expected_tables = ['users', 'updates', 'read_logs', 'activity_logs', 'archived_updates', 'archived_sop_summaries', 'archived_lessons_learned', 'sop_summaries', 'lessons_learned']

                logger.info(f"Existing tables: {existing_tables}")
                logger.info(f"Expected tables: {expected_tables}")

                missing_tables = [table for table in expected_tables if table not in existing_tables]
                if missing_tables:
                    logger.warning(f"Missing tables: {missing_tables}")
                    logger.info("Tables will be created automatically by SQLAlchemy")
                else:
                    logger.info("All expected tables are present")

                # Test a simple query on users table with connection health check
                try:
                    user_count = User.query.count()
                    logger.info(f"Users table has {user_count} records")
                except Exception as query_e:
                    logger.warning(f"User count query failed, but continuing: {query_e}")
            else:
                logger.info("Skipping table inspection in development mode")

            db_verification_success = True

        except Exception as e:
            logger.error(f"❌ Database verification failed: {e}")
            if os.getenv("FLASK_ENV") == "development":
                print(f"❌ Database verification failed: {e}")

            # Enhanced cleanup on error
            try:
                db.session.rollback()
                db.session.close()
                # Force engine disposal to clear any problematic connections
                db.engine.dispose()
                logger.info("Database engine disposed due to verification failure")
            except Exception as cleanup_e:
                logger.error(f"Error during cleanup: {cleanup_e}")

        # If database verification failed, try to create tables
        if not db_verification_success:
            logger.warning("❌ Database verification failed - attempting to create tables")
            try:
                # Force table creation
                db.create_all()
                logger.info("✅ Tables created successfully after verification failure")

                # Skip table inspection in development mode to avoid engine creation issues
                if os.getenv("FLASK_ENV") == "development":
                    logger.info("✅ Skipping table inspection in development mode")
                    db_verification_success = True
                else:
                    # Re-verify after table creation
                    inspector = inspect(db.engine)
                    existing_tables = inspector.get_table_names()
                    expected_tables = ['users', 'updates', 'read_logs', 'activity_logs', 'archived_updates', 'archived_sop_summaries', 'archived_lessons_learned', 'sop_summaries', 'lessons_learned']

                    missing_tables = [table for table in expected_tables if table not in existing_tables]
                    if missing_tables:
                        logger.error(f"❌ Still missing tables after creation attempt: {missing_tables}")
                        raise RuntimeError(f"Database table creation failed - missing: {missing_tables}")
                    else:
                        logger.info("✅ All tables verified after creation")
                        db_verification_success = True

            except Exception as create_e:
                logger.error(f"❌ Failed to create tables: {create_e}")
                if os.getenv("FLASK_ENV") == "production":
                    logger.warning("⚠️ Continuing with app startup despite table creation failure (production mode)")
                else:
                    raise RuntimeError(f"Database table creation failed: {create_e}")

    # Activity Logging Helper
    def log_activity(action, entity_type, entity_id, entity_title=None, details=None):
        """Log user activity for audit trail"""
        try:
            user_id = session.get("user_id")

            # Get client IP address
            client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'Unknown'))
            if ',' in client_ip:
                client_ip = client_ip.split(',')[0].strip()

            # Get user agent
            user_agent = request.headers.get('User-Agent', 'Unknown')

            activity = ActivityLog(
                user_id=user_id,
                action=action,
                entity_type=entity_type,
                entity_id=str(entity_id),
                entity_title=entity_title,
                timestamp=now_utc(),
                ip_address=client_ip,
                user_agent=user_agent,
                details=details
            )

            db.session.add(activity)
            db.session.commit()
        except Exception as e:
            # Don't let activity logging break the main functionality
            if os.getenv("FLASK_ENV") == "development":
                print(f"Activity logging failed: {e}")

    # Archive Helper Functions
    def archive_update(update):
        """Archive an update before deletion"""
        try:
            # Validate update object
            if not update or not update.id:
                raise ValueError("Invalid update object")

            # Get archiving user
            user_id = session.get("user_id")
            if not user_id:
                app.logger.warning("Archiving update without user context")

            # Check if already archived
            existing_archive = ArchivedUpdate.query.get(update.id)
            if existing_archive:
                app.logger.warning(f"Update {update.id} already archived")
                return True

            # Create archive record
            archived_update = ArchivedUpdate(
                id=update.id,
                name=update.name,
                process=update.process,
                message=update.message,
                timestamp=update.timestamp,
                archived_at=now_utc(),
                archived_by=user_id
            )

            # Save the archive
            db.session.add(archived_update)
            db.session.commit()
            
            app.logger.info(f"Successfully archived update {update.id}")
            return True
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to archive update {update.id if update and hasattr(update, 'id') else 'unknown'}: {str(e)}")
            return False

    def archive_sop(sop):
        """Archive a SOP before deletion"""
        try:
            # Validate SOP object
            if not sop or not sop.id:
                raise ValueError("Invalid SOP object")

            # Get archiving user
            user_id = session.get("user_id")
            if not user_id:
                app.logger.warning("Archiving SOP without user context")

            # Check if already archived
            existing_archive = ArchivedSOPSummary.query.get(sop.id)
            if existing_archive:
                app.logger.warning(f"SOP {sop.id} already archived")
                return True

            # Create archive record
            archived_sop = ArchivedSOPSummary(
                id=sop.id,
                title=sop.title,
                summary_text=sop.summary_text,
                department=sop.department,
                tags=sop.tags,
                created_at=sop.created_at,
                archived_at=now_utc(),
                archived_by=user_id
            )

            # Save the archive
            db.session.add(archived_sop)
            db.session.commit()
            
            app.logger.info(f"Successfully archived SOP {sop.id}")
            return True
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to archive SOP {sop.id if sop and hasattr(sop, 'id') else 'unknown'}: {str(e)}")
            return False

    def archive_lesson(lesson):
        """Archive a lesson learned before deletion"""
        try:
            # Validate lesson object
            if not lesson or not lesson.id:
                raise ValueError("Invalid lesson object")

            # Get archiving user
            user_id = session.get("user_id")
            if not user_id:
                app.logger.warning("Archiving lesson without user context")

            # Check if already archived
            existing_archive = ArchivedLessonLearned.query.get(lesson.id)
            if existing_archive:
                app.logger.warning(f"Lesson {lesson.id} already archived")
                return True

            # Create archive record
            archived_lesson = ArchivedLessonLearned(
                id=lesson.id,
                title=lesson.title,
                content=lesson.content,
                summary=lesson.summary,
                author=lesson.author,
                department=lesson.department,
                tags=lesson.tags,
                created_at=lesson.created_at,
                archived_at=now_utc(),
                archived_by=user_id
            )

            # Save the archive
            db.session.add(archived_lesson)
            db.session.commit()
            
            app.logger.info(f"Successfully archived lesson {lesson.id}")
            return True
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to archive lesson {lesson.id if lesson and hasattr(lesson, 'id') else 'unknown'}: {str(e)}")
            return False

    # Auth Helpers
    def login_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get("user_id"):
                flash("🔒 Please log in to continue.")
                return redirect(url_for("login", next=request.endpoint))
            return f(*args, **kwargs)
        return decorated

    @app.context_processor
    def inject_current_user():
        """Inject current user into template context with caching"""
        user_id = session.get("user_id")
        if user_id:
            # Use get() instead of filter_by().first() for better performance
            user = User.query.get(user_id)
            return dict(current_user=user)
        return dict(current_user=None)

    @app.context_processor
    def inject_user_role_info():
        """
        Inject user role information into templates for conditional rendering.
        """
        user_id = session.get("user_id")
        if user_id:
            role_info = get_cached_user_role(user_id)
            return dict(user_role=role_info)
        return dict(user_role={'is_admin': False, 'is_editor': False, 'is_writer': False, 'is_deleter': False, 'is_exporter': False})

    @app.context_processor
    def inject_now_utc():
        """Make now_utc function available in templates"""
        return dict(now_utc=now_utc)

    @app.template_filter('to_ist')
    def to_ist_filter(utc_datetime):
        """Convert UTC datetime to IST for display"""
        return format_ist(utc_datetime, '%Y-%m-%d %H:%M')

    @app.template_filter('format_datetime')
    def format_datetime_filter(dt, format='%Y-%m-%d %H:%M'):
        """Format datetime with timezone conversion"""
        return format_ist(dt, format)

    @app.template_filter('format_backup_timestamp')
    def format_backup_timestamp_filter(iso_string):
        """Convert ISO timestamp string to IST for backup display"""
        if not iso_string:
            return 'N/A'
        try:
            # Parse the ISO string to datetime
            from datetime import datetime
            # Handle different ISO formats
            if iso_string.endswith('Z'):
                iso_string = iso_string.replace('Z', '+00:00')
            elif '+' not in iso_string and 'T' in iso_string:
                # Assume UTC if no timezone info
                iso_string = iso_string + '+00:00'

            dt = datetime.fromisoformat(iso_string)
            # Convert to IST and format
            return format_ist(dt, '%Y-%m-%d %H:%M:%S')
        except (ValueError, AttributeError, TypeError):
            # Fallback to original string manipulation if parsing fails
            try:
                return iso_string[:19].replace('T', ' ') + ' (UTC)'
            except:
                return str(iso_string)

    @app.context_processor
    def built_assets_available():
        """Expose a small flag to templates indicating whether built CSS exists.

        Templates can use `built_css` to prefer a local built stylesheet
        (e.g. `static/dist/styles.css`) when present, otherwise fall back
        to a CDN link. This keeps CI builds optional and safe for local dev.
        """
        try:
            static_path = os.path.join(app.root_path, "static", "dist", "styles.css")
            is_prod = str(app.config.get('ENV', '')).lower() == 'production' or os.getenv('FLASK_ENV', '').lower() == 'production' or os.getenv('ENV', '').lower() == 'production'
            return dict(built_css=os.path.exists(static_path), is_production=is_prod)
        except Exception:
            return dict(built_css=False, is_production=False)

    # Routes
    @app.route("/health")
    @performance_logger
    def health():
        """Optimized health check endpoint with performance monitoring"""
        try:
            # Comprehensive database readiness check
            from database import ensure_database_ready
            if not ensure_database_ready():
                return jsonify({
                    "status": "error",
                    "message": "Database not ready",
                    "timestamp": now_utc().isoformat()
                }), 503

            # Test database connection efficiently
            db.session.execute(text("SELECT 1"))

            # Memory usage monitoring for free tier
            import psutil
            import os
            process = psutil.Process(os.getpid())
            memory_info = process.memory_info()
            memory_mb = memory_info.rss / 1024 / 1024

            out = {
                "status": "ok",
                "db": "connected",
                "database_type": db.engine.url.drivername,
                "memory_usage_mb": round(memory_mb, 2),
                "timestamp": now_utc().isoformat()
            }

            # Log memory usage for monitoring
            if memory_mb > 400:  # Log if memory usage is high for 512MB tier
                logger.warning(".2f"
                              f"memory_percent={(memory_mb/512)*100:.1f}")

            return jsonify(out), 200
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": str(e),
                "timestamp": now_utc().isoformat()
            }), 500

    # Metrics functionality removed for light version

    # Health alert functionality removed for light version

    @app.route("/lessons_learned/view/<int:lesson_id>")
    @login_required
    def view_lesson_learned(lesson_id):
        lesson = LessonLearned.query.get(lesson_id)
        if not lesson:
            flash("Lesson Learned not found.")
            return redirect(url_for("list_lessons_learned"))
        return render_template("view_lesson_learned.html", lesson=lesson)

    @app.route("/search")
    @performance_logger
    def search():
        query = request.args.get("q", "").strip()
        results = []

        # Get filter parameters
        category = request.args.get("category", "")
        process = request.args.get("process", "")
        department = request.args.get("department", "")
        date_range = request.args.get("date_range", "")

        # Prepare filters for template
        filters = {
            "category": category,
            "process": process,
            "department": department,
            "date_range": date_range
        }

        # Available options for filters
        available_processes = ["ABC", "XYZ", "AB"]
        available_departments = ["ABC", "XYZ", "AB"]  # Fixed to match process options

        if query:
            # Limit search results to prevent performance issues
            limit_per_category = 20

            # Case-insensitive search
            query_filter = f"%{query}%"

            # Search Updates (if no category filter or category is 'updates')
            if not category or category == "updates":
                updates_query = Update.query.filter(
                    or_(
                        Update.message.ilike(query_filter),
                        Update.name.ilike(query_filter),
                        Update.process.ilike(query_filter)
                    )
                )

                # Apply process filter
                if process:
                    updates_query = updates_query.filter(Update.process.ilike(f"%{process}%"))

                updates_rows = updates_query.order_by(Update.timestamp.desc()).limit(limit_per_category).all()

                for upd in updates_rows:
                    results.append({
                        "id": upd.id,
                        "title": f"{upd.process} - {upd.name}",
                        "content": upd.message[:200] + ("..." if len(upd.message) > 200 else ""),
                        "type": "update",
                        "url": url_for("view_update", update_id=upd.id),
                        "author": upd.name,
                        "created_at": upd.timestamp,
                        "process": upd.process
                    })

            # Search SOP Summaries (if no category filter or category is 'sops')
            if not category or category == "sops":
                sops_query = SOPSummary.query.filter(
                    or_(
                        SOPSummary.title.ilike(query_filter),
                        SOPSummary.summary_text.ilike(query_filter)
                    )
                )

                # Apply department filter
                if department:
                    sops_query = sops_query.filter(SOPSummary.department.ilike(f"%{department}%"))

                sops_rows = sops_query.order_by(SOPSummary.created_at.desc()).limit(limit_per_category).all()

                for sop in sops_rows:
                    results.append({
                        "id": sop.id,
                        "title": sop.title,
                        "content": sop.summary_text[:200] + ("..." if len(sop.summary_text) > 200 else ""),
                        "type": "sop",
                        "url": url_for("view_sop_summary", summary_id=sop.id),
                        "created_at": sop.created_at,
                        "tags": sop.tags or []
                    })

            # Search Lessons Learned (if no category filter or category is 'lessons')
            if not category or category == "lessons":
                lessons_query = LessonLearned.query.filter(
                    or_(
                        LessonLearned.title.ilike(query_filter),
                        LessonLearned.content.ilike(query_filter),
                        LessonLearned.summary.ilike(query_filter)
                    )
                )

                # Apply department filter
                if department:
                    lessons_query = lessons_query.filter(LessonLearned.department.ilike(f"%{department}%"))

                lessons_rows = lessons_query.order_by(LessonLearned.created_at.desc()).limit(limit_per_category).all()

                for lesson in lessons_rows:
                    results.append({
                        "id": lesson.id,
                        "title": lesson.title,
                        "content": (lesson.summary or lesson.content or "")[:200] + ("..." if len(lesson.summary or lesson.content or "") > 200 else ""),
                        "type": "lesson",
                        "url": url_for("view_lesson_learned", lesson_id=lesson.id),
                        "author": lesson.author,
                        "created_at": lesson.created_at,
                        "tags": lesson.tags or []
                    })

        return render_template("search_results.html",
                              query=query,
                              results=results,
                              filters=filters,
                              available_processes=available_processes,
                              available_departments=available_departments)
        query = request.args.get("q", "").strip()
        results = []

        # Get filter parameters
        category = request.args.get("category", "")
        process = request.args.get("process", "")
        department = request.args.get("department", "")
        date_range = request.args.get("date_range", "")

        # Prepare filters for template
        filters = {
            "category": category,
            "process": process,
            "department": department,
            "date_range": date_range
        }

        # Available options for filters
        available_processes = ["ABC", "XYZ", "AB"]
        available_departments = ["ABC", "XYZ", "AB"]  # Fixed to match process options

        if query:
            # Case-insensitive search
            query_filter = f"%{query}%"

            # Search Updates (if no category filter or category is 'updates')
            if not category or category == "updates":
                updates_query = Update.query.filter(
                    or_(
                        Update.message.ilike(query_filter),
                        Update.name.ilike(query_filter),
                        Update.process.ilike(query_filter)
                    )
                )

                # Apply process filter
                if process:
                    updates_query = updates_query.filter(Update.process.ilike(f"%{process}%"))

                updates_rows = updates_query.order_by(Update.timestamp.desc()).all()

                for upd in updates_rows:
                    results.append({
                        "id": upd.id,
                        "title": f"{upd.process} - {upd.name}",
                        "content": upd.message[:200] + ("..." if len(upd.message) > 200 else ""),
                        "type": "update",
                        "url": url_for("view_update", update_id=upd.id),
                        "author": upd.name,
                        "created_at": upd.timestamp,
                        "process": upd.process
                    })

            # Search SOP Summaries (if no category filter or category is 'sops')
            if not category or category == "sops":
                sops_query = SOPSummary.query.filter(
                    or_(
                        SOPSummary.title.ilike(query_filter),
                        SOPSummary.summary_text.ilike(query_filter)
                    )
                )

                # Apply department filter
                if department:
                    sops_query = sops_query.filter(SOPSummary.department.ilike(f"%{department}%"))

                sops_rows = sops_query.order_by(SOPSummary.created_at.desc()).all()

                for sop in sops_rows:
                    results.append({
                        "id": sop.id,
                        "title": sop.title,
                        "content": sop.summary_text[:200] + ("..." if len(sop.summary_text) > 200 else ""),
                        "type": "sop",
                        "url": url_for("view_sop_summary", summary_id=sop.id),
                        "created_at": sop.created_at,
                        "tags": sop.tags or []
                    })

            # Search Lessons Learned (if no category filter or category is 'lessons')
            if not category or category == "lessons":
                lessons_query = LessonLearned.query.filter(
                    or_(
                        LessonLearned.title.ilike(query_filter),
                        LessonLearned.content.ilike(query_filter),
                        LessonLearned.summary.ilike(query_filter)
                    )
                )

                # Apply department filter
                if department:
                    lessons_query = lessons_query.filter(LessonLearned.department.ilike(f"%{department}%"))

                lessons_rows = lessons_query.order_by(LessonLearned.created_at.desc()).all()

                for lesson in lessons_rows:
                    results.append({
                        "id": lesson.id,
                        "title": lesson.title,
                        "content": (lesson.summary or lesson.content or "")[:200] + ("..." if len(lesson.summary or lesson.content or "") > 200 else ""),
                        "type": "lesson",
                        "url": url_for("view_lesson_learned", lesson_id=lesson.id),
                        "author": lesson.author,
                        "created_at": lesson.created_at,
                        "tags": lesson.tags or []
                    })

        return render_template("search_results.html",
                             query=query,
                             results=results,
                             filters=filters,
                             available_processes=available_processes,
                             available_departments=available_departments)



    @app.route("/")
    @performance_logger
    def home():
        try:
            # Limit the number of items loaded for better performance
            summaries = SOPSummary.query.order_by(SOPSummary.created_at.desc()).limit(10).all()
            lessons = LessonLearned.query.order_by(LessonLearned.created_at.desc()).limit(10).all()

            # Get latest updates for the home page with read counts
            latest_updates = Update.query.order_by(Update.timestamp.desc()).limit(5).all()
            updates_data = []

            for update in latest_updates:
                update_dict = update.to_dict()
                # Get read count efficiently
                read_count = db.session.query(func.count(ReadLog.id)).filter(
                    ReadLog.update_id == update.id
                ).scalar()
                update_dict['read_count'] = read_count
                update_dict['is_new'] = is_within_hours(update.timestamp, 24, now_utc())
                updates_data.append(update_dict)

            return render_template("home.html", app_name=app.config["APP_NAME"],
                                  summaries=summaries, lessons=lessons, updates=updates_data,
                                  excel_export_available=EXCEL_EXPORT_AVAILABLE)
        except Exception as e:
            # Return a basic template without updates if there's an error
            return render_template("home.html", app_name=app.config["APP_NAME"],
                                  summaries=[], lessons=[], updates=[],
                                  excel_export_available=EXCEL_EXPORT_AVAILABLE)

    @app.route("/updates")
    @performance_logger
    def show_updates():
        # Get filter parameters from query string
        selected_process = request.args.get("process", "")
        selected_department = request.args.get("department", "")
        sort = request.args.get("sort", "newest")
        page = int(request.args.get("page", 1))
        per_page = min(int(request.args.get("per_page", 50)), 100)  # Limit to 100 max

        # Calculate offset for pagination
        offset = (page - 1) * per_page

        # Optimized base query with pagination
        base_query = db.session.query(Update)

        # Apply process filter if specified
        if selected_process:
            base_query = base_query.filter(Update.process == selected_process)

        # Apply sorting
        if sort == "oldest":
            base_query = base_query.order_by(Update.timestamp.asc())
        elif sort == "process":
            base_query = base_query.order_by(Update.process.asc(), Update.timestamp.desc())
        elif sort == "author":
            base_query = base_query.order_by(Update.name.asc(), Update.timestamp.desc())
        else:  # newest (default)
            base_query = base_query.order_by(Update.timestamp.desc())

        # Get total count for pagination with optimized query
        total_updates = base_query.count()

        # Apply pagination
        paginated_updates = base_query.offset(offset).limit(per_page).all()

        # Get read counts efficiently using a single query
        update_ids = [upd.id for upd in paginated_updates]
        if update_ids:
            read_counts = dict(
                db.session.query(ReadLog.update_id, func.count(ReadLog.id))
                .filter(ReadLog.update_id.in_(update_ids))
                .group_by(ReadLog.update_id)
                .all()
            )
        else:
            read_counts = {}

        updates = []
        current_time = now_utc()
        for upd in paginated_updates:
            d = upd.to_dict()
            d['read_count'] = read_counts.get(upd.id, 0)
            d['is_new'] = is_within_hours(upd.timestamp, 24, current_time)
            d['timestamp_obj'] = upd.timestamp
            updates.append(d)

        # Get additional data for template more efficiently
        unique_authors = [row[0] for row in db.session.query(Update.name).filter(Update.name.isnot(None)).distinct().all()]
        processes = [row[0] for row in db.session.query(Update.process).filter(Update.process.isnot(None)).distinct().all()]

        # Calculate updates this week more efficiently
        week_ago = get_hours_ago(24 * 7)
        updates_this_week = Update.query.filter(Update.timestamp >= week_ago).count()

        # Get unique departments (consistent with other forms)
        departments = ["ABC", "XYZ", "AB"]

        # Calculate pagination info
        total_pages = (total_updates + per_page - 1) // per_page
        has_next = page < total_pages
        has_prev = page > 1

        return render_template("show.html",
                              app_name=app.config["APP_NAME"],
                              updates=updates,
                              unique_authors=unique_authors,
                              processes=processes,
                              departments=departments,
                              updates_this_week=updates_this_week,
                              selected_process=selected_process,
                              selected_department=selected_department,
                              sort=sort,
                              page=page,
                              per_page=per_page,
                              total_updates=total_updates,
                              total_pages=total_pages,
                              has_next=has_next,
                              has_prev=has_prev)

    @app.route("/post", methods=["GET", "POST"])
    @writer_required
    def post_update():
        processes = ["ABC", "XYZ", "AB"]

        if request.method == "POST":
            # Validate connection before critical operation
            from database import validate_connection_before_operation
            if not validate_connection_before_operation():
                flash("⚠️ Database connection issue. Please try again.")
                return redirect(url_for("post_update"))

            message = request.form.get("message", "").strip()
            selected_process = request.form.get("process")
            name = inject_current_user()["current_user"].display_name

            if not message or not selected_process:
                flash("⚠️ Message and process are required.")
                return redirect(url_for("post_update"))

            new_update = Update(
                id=uuid.uuid4().hex,
                name=name,
                process=selected_process,
                message=message,
                timestamp=now_utc(),
            )
            try:
                db.session.add(new_update)
                db.session.commit()
                flash("✅ Update posted.")

                # Log activity
                log_activity('created', 'update', new_update.id, f"Update: {message[:50]}...")

                # Broadcast the new update via Socket.IO
                try:
                    if os.getenv("FLASK_ENV") == "development":
                        logger.info(f"🔄 Broadcasting new update via Socket.IO - ID: {new_update.id}")
                        print(f"🔄 Broadcasting new update via Socket.IO - ID: {new_update.id}")
                    from api.socketio import broadcast_update
                    update_data = {
                        'id': new_update.id,
                        'name': new_update.name,
                        'process': new_update.process,
                        'message': new_update.message,
                        'timestamp': new_update.timestamp.isoformat()
                    }
                    if os.getenv("FLASK_ENV") == "development":
                        logger.info(f"📦 Update data prepared: {update_data}")
                        print(f"📦 Update data prepared for broadcasting")
                    broadcast_update(update_data, selected_process)
                    if os.getenv("FLASK_ENV") == "development":
                        logger.info(f"✅ Broadcast function called successfully for update {new_update.id}")
                        print(f"✅ Broadcast function called successfully for update {new_update.id}")
                except Exception as e:
                    # Socket.IO broadcasting failure shouldn't break the posting
                    logger.error(f"❌ Socket.IO broadcast failed: {e}")
                    if os.getenv("FLASK_ENV") == "development":
                        print(f"❌ Socket.IO broadcast failed: {e}")

            except Exception as e:
                db.session.rollback()
                logger.error(f"Failed to post update: {e}")
                flash("⚠️ Failed to post update.")
            return redirect(url_for("show_updates"))

        return render_template("post.html", app_name=app.config["APP_NAME"], processes=processes)

    @app.route("/edit/<update_id>", methods=["GET", "POST"])
    @writer_required
    def edit_update(update_id):
        update = Update.query.get(update_id)
        current = inject_current_user()["current_user"]
        if not update or update.name != current.display_name:
            flash("🚫 Unauthorized or not found.")
            return redirect(url_for("show_updates"))

        if request.method == "POST":
            new_message = request.form.get("message", "").strip()
            if not new_message:
                flash("⚠️ Message cannot be empty.")
                return redirect(url_for("edit_update", update_id=update_id))
            update.message = new_message
            update.timestamp = now_utc()
            try:
                db.session.commit()
                flash("✏️ Update edited successfully.")

                # Log activity
                log_activity('edited', 'update', update.id, f"Update: {new_message[:50]}...")

            except Exception:
                db.session.rollback()
                flash("⚠️ Failed to edit update.")
            return redirect(url_for("show_updates"))

        return render_template("edit.html", app_name=app.config["APP_NAME"], update=update)

    @app.route("/view/<update_id>")
    def view_update(update_id):
        """View a specific update"""
        update = Update.query.get(update_id)
        if not update:
            flash("🚫 Update not found.")
            return redirect(url_for("show_updates"))
        
        # Get read count for this update
        read_count = db.session.query(func.count(ReadLog.id)).filter(
            ReadLog.update_id == update_id
        ).scalar()
        
        # Check if current user has read this update
        is_read = False
        if session.get("user_id"):
            read_log = ReadLog.query.filter_by(
                update_id=update_id,
                user_id=session["user_id"]
            ).first()
            is_read = read_log is not None
        
        return render_template("view_update.html", 
                             app_name=app.config["APP_NAME"], 
                             update=update, 
                             read_count=read_count,
                             is_read=is_read)

    @app.route("/delete/<update_id>", methods=["POST"])
    @login_required
    def delete_update(update_id):
        update = Update.query.get(update_id)
        current = inject_current_user()["current_user"]
        
        # Prepare response data
        response = {"success": False, "message": ""}
        
        if not update:
            response["message"] = "⚠️ Update not found."
            status_code = 404
        elif update.name.strip().lower() != current.display_name.strip().lower():
            response["message"] = "🚫 Not authorized to delete."
            status_code = 403
        else:
            # Capture details before deletion
            entity_title = f"Update: {update.message[:50]}..."

            try:
                # Archive the update before deletion
                if archive_update(update):
                    db.session.delete(update)
                    db.session.commit()
                    response["success"] = True
                    response["message"] = "✅ Update deleted and archived."
                    status_code = 200

                    # Log activity after successful deletion
                    log_activity('deleted', 'update', update_id, entity_title)
                else:
                    response["message"] = "⚠️ Failed to archive update. Deletion cancelled for safety."
                    status_code = 500

            except Exception as e:
                db.session.rollback()
                response["message"] = "❌ Deletion failed."
                status_code = 500
                logger.error(f"Failed to delete update: {e}")

        # Handle both AJAX and form submissions
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return jsonify(response), status_code
        else:
            flash(response["message"])
            return redirect(url_for("show_updates"))
            
    @app.route("/register", methods=["GET", "POST"])
    def register():
        if request.method == "POST":
            display_name = request.form["display_name"].strip()
            username = request.form["username"].strip().replace(" ", "_").lower()
            password = request.form["password"]

            if not re.match("^[A-Za-z0-9_]+$", username):
                flash("🚫 Username can only contain letters, numbers, and underscores.")
                return redirect(url_for("register"))

            if not username or not display_name or not password:
                flash("⚠️ All fields required.")
                return redirect(url_for("register"))

            if User.query.filter_by(username=username).first():
                flash("🚫 Username taken.")
                return redirect(url_for("register"))

            new_user = User(username=username, display_name=display_name)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()

            flash("✅ Registered! Please log in.")
            return redirect(url_for("login"))

        return render_template("register.html", app_name=app.config["APP_NAME"])

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form["username"].strip().replace(" ", "_").lower()
            password = request.form["password"]
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                session["user_id"] = user.id
                flash(f"👋 Welcome back, {user.display_name}!")

                # Handle redirect logic
                next_page = request.form.get('next') or request.args.get('next')

                # If user came from home page login button, redirect to home
                if next_page == 'home':
                    return redirect(url_for("home"))
                # If there's a next page specified, redirect there
                elif next_page:
                    # Check if it's a full URL (from role decorators)
                    if next_page.startswith('http') or next_page.startswith('/'):
                        # Validate that it's a safe redirect within our app
                        from urllib.parse import urlparse
                        parsed = urlparse(next_page)
                        if not parsed.netloc or parsed.netloc == request.host:
                            return redirect(next_page)

                    # Map of valid endpoints to their URL functions
                    valid_endpoints = {
                        'show_updates': 'show_updates',
                        'post_update': 'post_update',
                        'list_sop_summaries': 'list_sop_summaries',
                        'list_lessons_learned': 'list_lessons_learned',
                        'search': 'search',
                        'add_sop_summary': 'add_sop_summary',
                        'add_lesson_learned': 'add_lesson_learned',
                        'export_readlogs': 'export_readlogs'
                    }

                    # Handle search result redirects for SOPs and Lessons
                    if next_page == 'view_sop_summary':
                        return redirect(url_for('list_sop_summaries'))
                    elif next_page == 'view_lesson_learned':
                        return redirect(url_for('list_lessons_learned'))
                    elif next_page in valid_endpoints:
                        return redirect(url_for(valid_endpoints[next_page]))

                # Default fallback to home page instead of show_updates
                return redirect(url_for("home"))

            flash("🚫 Invalid credentials.")
            return redirect(url_for("login"))

        # For GET requests, preserve the next parameter
        next_page = request.args.get('next')
        return render_template("login.html", app_name=app.config["APP_NAME"], next_page=next_page)

    @app.route("/logout")
    def logout():
        session.pop("user_id", None)
        flash("👋 You’ve been logged out.")
        return redirect(url_for("home"))

    @app.route("/api/latest-update-time")
    def api_latest_update_time():
        """API endpoint to get the timestamp of the most recent update"""
        try:
            # Use proper database session management to prevent lock issues
            from database import db_session
            with db_session() as session:
                # Use optimized query for free tier - only select timestamp column
                latest_timestamp = session.query(Update.timestamp).order_by(Update.timestamp.desc()).first()

                if latest_timestamp:
                    # Ensure timezone is properly handled
                    timestamp = ensure_timezone(latest_timestamp.timestamp, UTC)
                    result = {
                        "latest_timestamp": timestamp.isoformat(),
                        "success": True
                    }
                else:
                    result = {
                        "latest_timestamp": None,
                        "success": True
                    }

                return jsonify(result)

        except Exception as e:
            logger.error(f"Error getting latest update time: {e}")
            return jsonify({
                "latest_timestamp": None,
                "success": False,
                "error": str(e)
            }), 500
    
    # New routes for SOP Summaries
    @app.route("/sop_summaries")
    @login_required
    @performance_logger
    def list_sop_summaries():
        page = int(request.args.get("page", 1))
        per_page = min(int(request.args.get("per_page", 20)), 50)  # Limit to 50 max
        offset = (page - 1) * per_page

        # Get paginated results
        paginated_sops = SOPSummary.query.order_by(SOPSummary.created_at.desc()).offset(offset).limit(per_page).all()

        # Get total count for pagination
        total_sops = SOPSummary.query.count()
        total_pages = (total_sops + per_page - 1) // per_page

        return render_template("sop_summaries.html",
                             summaries=paginated_sops,
                             page=page,
                             per_page=per_page,
                             total_sops=total_sops,
                             total_pages=total_pages,
                             has_next=page < total_pages,
                             has_prev=page > 1)

    @app.route("/sop_summaries/add", methods=["GET", "POST"])
    @writer_required
    def add_sop_summary():
        if request.method == "POST":
            title = request.form.get("title", "").strip()
            summary_text = request.form.get("summary_text", "").strip()
            department = request.form.get("department", "").strip()
            tags = request.form.get("tags", "").strip()

            if not title or not summary_text:
                flash("Title and Summary are required.")
                return redirect(url_for("add_sop_summary"))

            tags_list = [tag.strip() for tag in tags.split(",")] if tags else []

            sop = SOPSummary(
                title=title,
                summary_text=summary_text,
                department=department or None,
                tags=tags_list or None,
            )
            try:
                db.session.add(sop)
                db.session.commit()
                flash("SOP Summary added successfully.")

                # Log activity
                log_activity('created', 'sop', sop.id, title)

                # Broadcast notification for new SOP
                try:
                    from api.socketio import broadcast_notification
                    broadcast_notification({
                        'type': 'new_sop',
                        'title': 'New SOP Added',
                        'message': f'**NEW SOP:** {title}',
                        'timestamp': now_utc().isoformat()
                    })
                except Exception as e:
                    if os.getenv("FLASK_ENV") == "development":
                        print(f"Socket.IO broadcast failed: {e}")

                return redirect(url_for("list_sop_summaries"))
            except Exception as e:
                db.session.rollback()
                flash("Failed to add SOP Summary.")
                logger.error(f"Failed to add SOP Summary: {e}")
                return redirect(url_for("add_sop_summary"))

        return render_template("add_sop_summary.html")

    @app.route("/sop_summaries/edit/<int:sop_id>", methods=["GET", "POST"])
    @writer_required
    def edit_sop_summary(sop_id):
        sop = SOPSummary.query.get(sop_id)
        if not sop:
            flash("SOP Summary not found.")
            return redirect(url_for("list_sop_summaries"))

        if request.method == "POST":
            title = request.form.get("title", "").strip()
            summary_text = request.form.get("summary_text", "").strip()
            department = request.form.get("department", "").strip()
            tags = request.form.get("tags", "").strip()

            if not title or not summary_text:
                flash("Title and Summary are required.")
                return redirect(url_for("edit_sop_summary", sop_id=sop_id))

            tags_list = [tag.strip() for tag in tags.split(",")] if tags else []

            sop.title = title
            sop.summary_text = summary_text
            sop.department = department or None
            sop.tags = tags_list or None
            try:
                db.session.commit()
                flash("✅ SOP Summary updated successfully.")

                # Log activity
                log_activity('edited', 'sop', sop.id, title)

            except Exception as e:
                db.session.rollback()
                flash("❌ Failed to update SOP Summary.")
                logger.error(f"Failed to update SOP Summary: {e}")

            # Check if user came from view page or list page
            referrer = request.form.get('referrer') or request.referrer
            if referrer and 'sop_summaries/' in referrer and str(sop_id) in referrer:
                return redirect(url_for("view_sop_summary", summary_id=sop_id))
            else:
                return redirect(url_for("list_sop_summaries"))

        return render_template("edit_sop_summary.html", sop=sop)

    @app.route("/sop_summaries/delete/<int:sop_id>", methods=["POST"])
    @delete_required
    def delete_sop_summary(sop_id):
        sop = SOPSummary.query.get(sop_id)
        if not sop:
            flash("⚠️ SOP Summary not found.")
            return redirect(url_for("list_sop_summaries"))

        # Capture details before deletion
        entity_title = sop.title

        try:
            # Archive the SOP before deletion
            if archive_sop(sop):
                db.session.delete(sop)
                db.session.commit()
                flash("✅ SOP Summary deleted and archived.")

                # Log activity after successful deletion
                log_activity('deleted', 'sop', sop_id, entity_title)
            else:
                flash("⚠️ Failed to archive SOP. Deletion cancelled for safety.")

        except Exception as e:
            db.session.rollback()
            flash("❌ Failed to delete SOP Summary.")
            logger.error(f"Failed to delete SOP Summary: {e}")

        # Always redirect to list page after deletion since the item no longer exists
        return redirect(url_for("list_sop_summaries"))

    # New routes for Lessons Learned
    @app.route("/lessons_learned")
    @login_required
    @performance_logger
    def list_lessons_learned():
        page = int(request.args.get("page", 1))
        per_page = min(int(request.args.get("per_page", 20)), 50)  # Limit to 50 max
        offset = (page - 1) * per_page

        # Get paginated results
        paginated_lessons = LessonLearned.query.order_by(LessonLearned.created_at.desc()).offset(offset).limit(per_page).all()

        # Get total count for pagination
        total_lessons = LessonLearned.query.count()
        total_pages = (total_lessons + per_page - 1) // per_page

        return render_template("lessons_learned.html",
                             lessons=paginated_lessons,
                             page=page,
                             per_page=per_page,
                             total_lessons=total_lessons,
                             total_pages=total_pages,
                             has_next=page < total_pages,
                             has_prev=page > 1)

    @app.route("/lessons_learned/add", methods=["GET", "POST"])
    @writer_required
    def add_lesson_learned():
        if request.method == "POST":
            title = request.form.get("title", "").strip()
            content = request.form.get("content", "").strip()
            summary = request.form.get("summary", "").strip()
            author = request.form.get("author", "").strip()
            department = request.form.get("department", "").strip()
            tags = request.form.get("tags", "").strip()

            if not title or not content:
                flash("Title and Content are required.")
                return redirect(url_for("add_lesson_learned"))

            tags_list = [tag.strip() for tag in tags.split(",")] if tags else []

            lesson = LessonLearned(
                title=title,
                content=content,
                summary=summary or None,
                author=author or None,
                department=department or None,
                tags=tags_list or None,
            )
            try:
                db.session.add(lesson)
                db.session.commit()
                flash("Lesson Learned added successfully.")

                # Log activity
                log_activity('created', 'lesson', lesson.id, title)

                # Broadcast notification for new Lesson Learned
                try:
                    from api.socketio import broadcast_notification
                    broadcast_notification({
                        'type': 'new_lesson',
                        'title': 'New Lesson Learned',
                        'message': f'**NEW LESSON:** {title}',
                        'timestamp': now_utc().isoformat()
                    })
                except Exception as e:
                    if os.getenv("FLASK_ENV") == "development":
                        print(f"Socket.IO broadcast failed: {e}")

                return redirect(url_for("list_lessons_learned"))
            except Exception as e:
                db.session.rollback()
                flash("Failed to add Lesson Learned.")
                logger.error(f"Failed to add Lesson Learned: {e}")
                return redirect(url_for("add_lesson_learned"))

        return render_template("add_lesson_learned.html")

    @app.route("/lessons_learned/edit/<int:lesson_id>", methods=["GET", "POST"])
    @writer_required
    def edit_lesson_learned(lesson_id):
        lesson = LessonLearned.query.get(lesson_id)
        if not lesson:
            flash("Lesson Learned not found.")
            return redirect(url_for("list_lessons_learned"))

        if request.method == "POST":
            title = request.form.get("title", "").strip()
            content = request.form.get("content", "").strip()
            summary = request.form.get("summary", "").strip()
            author = request.form.get("author", "").strip()
            department = request.form.get("department", "").strip()
            tags = request.form.get("tags", "").strip()

            if not title or not content:
                flash("Title and Content are required.")
                return redirect(url_for("edit_lesson_learned", lesson_id=lesson_id))

            tags_list = [tag.strip() for tag in tags.split(",")] if tags else []

            lesson.title = title
            lesson.content = content
            lesson.summary = summary or None
            lesson.author = author or None
            lesson.department = department or None
            lesson.tags = tags_list or None

            try:
                db.session.commit()
                flash("✅ Lesson Learned updated successfully.")

                # Log activity
                log_activity('edited', 'lesson', lesson.id, title)

            except Exception as e:
                db.session.rollback()
                flash("❌ Failed to update Lesson Learned.")
                logger.error(f"Failed to update Lesson Learned: {e}")

            # Check if user came from view page or list page
            referrer = request.form.get('referrer') or request.referrer
            if referrer and 'lessons_learned/view/' in referrer and str(lesson_id) in referrer:
                return redirect(url_for("view_lesson_learned", lesson_id=lesson_id))
            else:
                return redirect(url_for("list_lessons_learned"))

        return render_template("edit_lesson_learned.html", lesson=lesson)

    @app.route("/lessons_learned/delete/<int:lesson_id>", methods=["POST"])
    @delete_required
    def delete_lesson_learned(lesson_id):
        lesson = LessonLearned.query.get(lesson_id)
        if not lesson:
            flash("Lesson Learned not found.")
            return redirect(url_for("list_lessons_learned"))

        # Capture details before deletion
        entity_title = lesson.title

        try:
            # Archive the lesson before deletion
            if archive_lesson(lesson):
                db.session.delete(lesson)
                db.session.commit()
                flash("✅ Lesson Learned deleted and archived.")

                # Log activity after successful deletion
                log_activity('deleted', 'lesson', lesson_id, entity_title)
            else:
                flash("⚠️ Failed to archive lesson. Deletion cancelled for safety.")

        except Exception as e:
            db.session.rollback()
            flash("❌ Failed to delete Lesson Learned.")

        # Always redirect to list page after deletion since the item no longer exists
        return redirect(url_for("list_lessons_learned"))
    

        
    @app.route("/sop_summaries/<int:summary_id>")
    @login_required
    def view_sop_summary(summary_id):
        summary = SOPSummary.query.get(summary_id)
        if not summary:
            flash("\u26a0\ufe0f SOP Summary not found.")
            return redirect(url_for("list_sop_summaries"))

        return render_template(
            "view_sop_summary.html",
            summary=summary,
            app_name=current_app.config["APP_NAME"]
        )



    @app.route("/api/check-update/<update_id>")
    def check_update(update_id):
        """Check if an update exists and is accessible."""
        try:
            update = Update.query.get(update_id)
            if not update:
                return jsonify({
                    'error': 'Not Found',
                    'message': 'Update not found'
                }), 404
            return jsonify({
                'exists': True,
                'id': update.id,
                'status': 'active'
            })
        except Exception as e:
            logger.error(f"Error checking update {update_id}: {str(e)}")
            return jsonify({
                'error': 'Internal Server Error',
                'message': 'Unable to check update status'
            }), 500

    @app.route("/api/recent-updates")
    def recent_updates():
        """Get updates from the past 24 hours for the bell icon banner."""
        try:
            # Calculate 24 hours ago
            twenty_four_hours_ago = get_hours_ago(24)

            # Use more efficient query with explicit column selection for free tier optimization
            recent_updates = db.session.query(
                Update.id,
                Update.name,
                Update.process,
                Update.message,
                Update.timestamp
            ).filter(
                Update.timestamp >= twenty_four_hours_ago
            ).order_by(Update.timestamp.desc()).limit(10).all()

            if not recent_updates:
                return jsonify({
                    'updates': [],
                    'message': 'No recent updates found'
                })

            updates_data = []
            for update in recent_updates:
                # Ensure timezone is properly handled
                timestamp = ensure_timezone(update.timestamp, UTC)

                updates_data.append({
                    "id": update.id,
                    "name": update.name,
                    "process": update.process,
                    "message": update.message,
                    "timestamp": timestamp.isoformat()
                })

            return jsonify({
                "success": True,
                "updates": updates_data
            })
        except Exception as e:
            logger.error(f"Error in recent-updates: {str(e)}")
            return jsonify({
                "success": False,
                "error": str(e)
            }), 500

    # Excel export functionality for read logs with multiple sheets
    @app.route("/export_readlogs")
    @export_required
    @performance_logger
    def export_readlogs():
        """Enhanced Excel export with read logs, activity logs, analytics, and metrics in separate sheets"""
        logger.info("Export read logs function called")
        download = request.args.get('download', 'false').lower() == 'true'
        logger.info(f"Download parameter: {download}")

        if not download:
            # Show the export page
            logger.info("Showing export page (not downloading)")
            return render_template("export_readlogs.html", app_name=app.config["APP_NAME"])

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            # Create Excel workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Sheet 1: Read Logs
            ws_readlogs = wb.create_sheet("Read Logs")
            ws_readlogs.append([
                'Read ID', 'Update ID', 'Reader Type', 'Reader Name', 'Timestamp (IST)',
                'IP Address', 'User Agent', 'Update Title', 'Process', 'Update Timestamp (IST)'
            ])

            # Style header row
            for col_num, cell in enumerate(ws_readlogs[1], 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Get read logs data
            with app.app_context():
                read_logs = db.session.query(
                    ReadLog.id,
                    ReadLog.update_id,
                    ReadLog.user_id,
                    ReadLog.guest_name,
                    ReadLog.timestamp,
                    ReadLog.ip_address,
                    ReadLog.user_agent,
                    Update.name.label('update_name'),
                    Update.process,
                    Update.timestamp.label('update_timestamp')
                ).join(
                    Update, ReadLog.update_id == Update.id
                ).order_by(
                    ReadLog.timestamp.desc()
                ).limit(10000).all()  # Limit for performance

            for log in read_logs:
                try:
                    reader_type = 'Registered' if log.user_id else 'Guest'
                    reader_name = log.guest_name if not log.user_id else 'N/A'
                    ist_timestamp = format_ist(log.timestamp, '%Y-%m-%d %H:%M:%S')
                    update_ist_timestamp = format_ist(log.update_timestamp, '%Y-%m-%d %H:%M:%S')

                    ws_readlogs.append([
                        log.id,
                        log.update_id,
                        reader_type,
                        reader_name,
                        ist_timestamp,
                        log.ip_address or '',
                        (log.user_agent or '')[:100],
                        log.update_name,
                        log.process,
                        update_ist_timestamp
                    ])
                except Exception as row_error:
                    logger.error(f"Error processing read log entry {log.id if hasattr(log, 'id') else 'unknown'}: {str(row_error)}")
                    continue

            # Sheet 2: Activity Logs
            ws_activity = wb.create_sheet("Activity Logs")
            ws_activity.append([
                'Activity ID', 'User', 'Action', 'Entity Type', 'Entity Title',
                'Timestamp (IST)', 'IP Address', 'User Agent', 'Details'
            ])

            # Style header row
            for col_num, cell in enumerate(ws_activity[1], 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Get activity logs
            with app.app_context():
                activity_logs = db.session.query(
                    ActivityLog.id,
                    ActivityLog.user_id,
                    ActivityLog.action,
                    ActivityLog.entity_type,
                    ActivityLog.entity_title,
                    ActivityLog.timestamp,
                    ActivityLog.ip_address,
                    ActivityLog.user_agent,
                    ActivityLog.details,
                    User.display_name.label('user_name')
                ).outerjoin(
                    User, ActivityLog.user_id == User.id
                ).order_by(
                    ActivityLog.timestamp.desc()
                ).limit(5000).all()  # Limit for performance

            for log in activity_logs:
                try:
                    ist_timestamp = format_ist(log.timestamp, '%Y-%m-%d %H:%M:%S')
                    user_name = log.user_name if log.user_name else 'System'

                    ws_activity.append([
                        log.id,
                        user_name,
                        log.action,
                        log.entity_type,
                        log.entity_title or '',
                        ist_timestamp,
                        log.ip_address or '',
                        (log.user_agent or '')[:100],
                        log.details or ''
                    ])
                except Exception as row_error:
                    logger.error(f"Error processing activity log entry {log.id}: {str(row_error)}")
                    continue

            # Sheet 3: Summary Analytics
            ws_analytics = wb.create_sheet("Summary Analytics")
            ws_analytics.append(['Metric', 'Value'])

            # Style header row
            for col_num, cell in enumerate(ws_analytics[1], 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            with app.app_context():
                # Total reads
                total_reads = db.session.query(func.count(ReadLog.id)).scalar()

                # Unique readers (registered users)
                unique_registered = db.session.query(func.count(func.distinct(ReadLog.user_id))).filter(
                    ReadLog.user_id.isnot(None)
                ).scalar()

                # Unique guest readers
                unique_guests = db.session.query(func.count(func.distinct(ReadLog.guest_name))).filter(
                    ReadLog.user_id.is_(None)
                ).scalar()

                # Total registered vs guest reads
                registered_reads = db.session.query(func.count(ReadLog.id)).filter(
                    ReadLog.user_id.isnot(None)
                ).scalar()

                guest_reads = db.session.query(func.count(ReadLog.id)).filter(
                    ReadLog.user_id.is_(None)
                ).scalar()

                # Total updates
                total_updates = db.session.query(func.count(Update.id)).scalar()

            analytics_data = [
                ['Total Reads', total_reads or 0],
                ['Unique Registered Readers', unique_registered or 0],
                ['Unique Guest Readers', unique_guests or 0],
                ['Registered User Reads', registered_reads or 0],
                ['Guest User Reads', guest_reads or 0],
                ['Total Updates', total_updates or 0]
            ]

            for row in analytics_data:
                ws_analytics.append(row)

            # Sheet 4: Top Performers
            ws_performers = wb.create_sheet("Top Performers")

            # Most active readers section
            ws_performers.append(['Most Active Readers'])
            ws_performers.append(['Reader Name', 'Reader Type', 'Total Reads'])

            # Style headers
            for col_num, cell in enumerate(ws_performers[2], 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            with app.app_context():
                # Top registered readers
                top_registered = db.session.query(
                    User.display_name,
                    func.count(ReadLog.id).label('read_count')
                ).join(
                    ReadLog, User.id == ReadLog.user_id
                ).group_by(
                    User.id, User.display_name
                ).order_by(
                    func.count(ReadLog.id).desc()
                ).limit(10).all()

                # Top guest readers
                top_guests = db.session.query(
                    ReadLog.guest_name,
                    func.count(ReadLog.id).label('read_count')
                ).filter(
                    ReadLog.user_id.is_(None)
                ).group_by(
                    ReadLog.guest_name
                ).order_by(
                    func.count(ReadLog.id).desc()
                ).limit(10).all()

            for reader, count in top_registered:
                ws_performers.append([reader, 'Registered', count])

            for reader, count in top_guests:
                ws_performers.append([reader, 'Guest', count])

            # Most popular updates section
            ws_performers.append([])
            ws_performers.append(['Most Popular Updates'])
            ws_performers.append(['Update Title', 'Process', 'Total Reads'])

            with app.app_context():
                top_updates = db.session.query(
                    Update.name,
                    Update.process,
                    func.count(ReadLog.id).label('read_count')
                ).outerjoin(
                    ReadLog, Update.id == ReadLog.update_id
                ).group_by(
                    Update.id, Update.name, Update.process
                ).order_by(
                    func.count(ReadLog.id).desc()
                ).limit(10).all()

            for title, process, count in top_updates:
                ws_performers.append([title, process, count])

            # Sheet 5: Engagement Metrics by Process
            ws_engagement = wb.create_sheet("Engagement by Process")
            ws_engagement.append(['Process', 'Total Updates', 'Total Reads', 'Unique Readers', 'Avg Reads per Update'])

            # Style header row
            for col_num, cell in enumerate(ws_engagement[1], 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            with app.app_context():
                # Get process metrics with proper unique reader counting
                process_metrics = db.session.query(
                    Update.process,
                    func.count(func.distinct(Update.id)).label('update_count'),
                    func.count(ReadLog.id).label('read_count')
                ).outerjoin(
                    ReadLog, Update.id == ReadLog.update_id
                ).group_by(
                    Update.process
                ).order_by(
                    Update.process
                ).all()

                # Get unique readers per process
                process_unique_readers = {}
                for process_name, _, _ in process_metrics:
                    # Count unique registered users for this process
                    registered_count = db.session.query(
                        func.count(func.distinct(ReadLog.user_id))
                    ).join(
                        Update, ReadLog.update_id == Update.id
                    ).filter(
                        Update.process == process_name,
                        ReadLog.user_id.isnot(None)
                    ).scalar()

                    # Count unique guest users for this process
                    guest_count = db.session.query(
                        func.count(func.distinct(ReadLog.guest_name))
                    ).join(
                        Update, ReadLog.update_id == Update.id
                    ).filter(
                        Update.process == process_name,
                        ReadLog.user_id.is_(None),
                        ReadLog.guest_name.isnot(None)
                    ).scalar()

                    process_unique_readers[process_name] = (registered_count or 0) + (guest_count or 0)

            for process, update_count, read_count in process_metrics:
                unique_readers = process_unique_readers.get(process, 0)
                avg_reads = round(read_count / update_count, 2) if update_count > 0 else 0
                ws_engagement.append([process, update_count, read_count, unique_readers, avg_reads])

            # Auto-adjust column widths for all sheets
            for ws in [ws_readlogs, ws_activity, ws_analytics, ws_performers, ws_engagement]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Generate timestamp for filename
            timestamp_str = format_ist(now_utc(), '%Y%m%d_%H%M%S')

            # Save workbook to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Return Excel response
            response = Response(
                excel_buffer.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=readlogs_export_{timestamp_str}.xlsx',
                    'Cache-Control': 'no-cache'
                }
            )
            return response

        except ImportError:
            logger.error("openpyxl not available, falling back to CSV export")
            flash("Excel export library not available. Please install required dependencies.", "error")
            return redirect(url_for('export_readlogs'))
        except Exception as e:
            logger.error(f"Error generating Excel export: {e}")
            flash("Error generating export file. Please try again.", "error")
            return redirect(url_for('export_readlogs'))

    @app.route("/reset_activity_logs", methods=["POST"])
    @admin_required
    def reset_activity_logs():
        """Reset all activity logs - admin only with password confirmation"""
        try:
            # Get admin password from form
            admin_password = request.form.get('admin_password', '').strip()

            if not admin_password:
                flash("Admin password is required.", "error")
                return redirect(url_for('export_readlogs'))

            # Verify admin password
            current_user = User.query.get(session.get("user_id"))
            if not current_user or not current_user.check_password(admin_password):
                flash("Invalid admin password.", "error")
                return redirect(url_for('export_readlogs'))

            # Get count before deletion for logging
            count_before = ActivityLog.query.count()

            # Delete all activity logs
            ActivityLog.query.delete()

            # Commit the changes
            db.session.commit()

            # Log this action (this will create a new activity log entry)
            log_activity('reset', 'activity_logs', 'all', f'Reset {count_before} activity log records')

            flash(f"Successfully reset {count_before} activity log records.", "success")

        except Exception as e:
            db.session.rollback()
            logger.error(f"Failed to reset activity logs: {e}")
            flash("Failed to reset activity logs. Please try again.", "error")

        return redirect(url_for('export_readlogs'))

    # Database Backup Routes
    @app.route("/backup")
    @admin_required
    @performance_logger
    def backup_page():
        """Show database backup management page"""
        try:
            from backup_system import DatabaseBackupSystem
            backup_system = DatabaseBackupSystem()
            backups = backup_system.list_backups()

            return render_template("backup.html",
                                 app_name=app.config["APP_NAME"],
                                 backups=backups)
        except ImportError as e:
            logger.error(f"Failed to import backup system: {e}")
            flash("Backup system module not found.", "error")
            return redirect(url_for("home"))
        except Exception as e:
            logger.error(f"Error loading backup page: {e}")
            flash("Error loading backup page. Please try again later.", "error")
            return redirect(url_for("home"))

    @app.route("/backup/create", methods=["POST"])
    @admin_required
    @performance_logger
    def create_backup():
        """Create a new database backup"""
        try:
            from backup_system import DatabaseBackupSystem
            backup_system = DatabaseBackupSystem()

            backup_type = request.form.get("backup_type", "manual")
            backup_path = backup_system.create_backup(backup_type)

            if backup_path:
                flash(f"✅ Database backup created successfully: {backup_path}", "success")
                # Log the backup creation
                log_activity('created', 'backup', 'system', f'Database backup: {backup_type}')
            else:
                flash("❌ Failed to create database backup.", "error")

        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            flash("❌ Error creating backup.", "error")

        return redirect(url_for('backup_page'))

    @app.route("/backup/download/<filename>")
    @admin_required
    @performance_logger
    def download_backup(filename):
        """Download a backup file"""
        try:
            from backup_system import DatabaseBackupSystem
            backup_system = DatabaseBackupSystem()

            # Find the backup file
            backups = backup_system.list_backups()
            backup_info = next((b for b in backups if b['filename'] == filename), None)

            if not backup_info:
                flash("Backup file not found.", "error")
                return redirect(url_for('backup_page'))

            backup_path = backup_info['path']
            json_path = f"{backup_path}.json"

            if not os.path.exists(json_path):
                flash("Backup file not found on disk.", "error")
                return redirect(url_for('backup_page'))

            # Log the download
            log_activity('downloaded', 'backup', filename, f'Backup file downloaded: {filename}')

            return send_file(json_path,
                           as_attachment=True,
                           download_name=f"{filename}.json",
                           mimetype='application/json')

        except Exception as e:
            logger.error(f"Error downloading backup: {e}")
            flash("Error downloading backup file.", "error")
            return redirect(url_for('backup_page'))

    @app.route("/backup/delete/<filename>", methods=["POST"])
    @admin_required
    @performance_logger
    def delete_backup(filename):
        """Delete a backup file"""
        try:
            from backup_system import DatabaseBackupSystem
            backup_system = DatabaseBackupSystem()

            # Find the backup file
            backups = backup_system.list_backups()
            backup_info = next((b for b in backups if b['filename'] == filename), None)

            if not backup_info:
                flash("Backup file not found.", "error")
                return redirect(url_for('backup_page'))

            backup_path = backup_info['path']
            json_path = f"{backup_path}.json"

            if os.path.exists(json_path):
                os.remove(json_path)
                flash(f"✅ Backup file '{filename}' deleted successfully.", "success")
                # Log the deletion
                log_activity('deleted', 'backup', filename, f'Backup file deleted: {filename}')
            else:
                flash("Backup file not found on disk.", "error")

        except Exception as e:
            logger.error(f"Error deleting backup: {e}")
            flash("Error deleting backup file.", "error")

        return redirect(url_for('backup_page'))

    @app.route("/backup/cleanup", methods=["POST"])
    @admin_required
    @performance_logger
    def cleanup_backups():
        """Clean up old backup files"""
        try:
            from backup_system import DatabaseBackupSystem
            backup_system = DatabaseBackupSystem()

            keep_days = int(request.form.get("keep_days", 30))
            deleted_count = backup_system.cleanup_old_backups(keep_days)

            if deleted_count > 0:
                flash(f"✅ Cleaned up {deleted_count} old backup files.", "success")
                # Log the cleanup
                log_activity('cleanup', 'backup', 'system', f'Cleaned up {deleted_count} old backup files')
            else:
                flash("No old backup files to clean up.", "info")

        except Exception as e:
            logger.error(f"Error cleaning up backups: {e}")
            flash("Error cleaning up backup files.", "error")

        return redirect(url_for('backup_page'))

    # Additional functionality removed for light version

    # Add cache control for static files to prevent 304 caching issues
    @app.after_request
    def add_cache_control(response):
        if request.path.startswith('/static/'):
            # Don't cache JS and CSS files in development
            if app.config.get('DEBUG', False) or os.getenv('FLASK_ENV') != 'production':
                if request.path.endswith(('.js', '.css')):
                    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response.headers['Pragma'] = 'no-cache'
                    response.headers['Expires'] = '0'
            else:
                # In production, cache static files aggressively
                if request.path.endswith(('.js', '.css', '.png', '.jpg', '.jpeg', '.gif', '.ico', '.svg')):
                    response.headers['Cache-Control'] = 'public, max-age=31536000'  # 1 year
                    response.headers['Expires'] = (datetime.utcnow() + timedelta(days=365)).strftime('%a, %d %b %Y %H:%M:%S GMT')

        # Socket.IO CORS is handled by the Socket.IO extension, don't override here

        # Add Vercel-specific headers (only essential ones in production)
        is_production = os.getenv("FLASK_ENV") == "production"
        is_development = os.getenv("FLASK_ENV") == "development" or not is_production

        # Essential headers for all environments
        response.headers['X-SocketIO-Support'] = 'enabled'
        response.headers['X-WebSocket-Support'] = 'enabled'
        response.headers['X-Transport-Support'] = 'websocket,polling'
        response.headers['X-Server-Version'] = 'LoopIn-v1.0'

        # Add debug headers only in development
        if is_development:
            response.headers['X-Debug-Mode'] = 'true'
            response.headers['X-Timestamp'] = now_utc().isoformat()
            response.headers['X-Status'] = 'healthy'
            response.headers['X-Connection-Status'] = 'ready'
            response.headers['X-SocketIO-Version'] = '4.7.2'
            response.headers['X-Debug-Info'] = 'Socket.IO debugging enabled'

        return response

    # Global error handlers to prevent worker crashes
    @app.errorhandler(404)
    def handle_404_error(error):
        """Handle 404 Not Found Error"""
        logger.warning(f"404 Not Found: {error}")
        return jsonify({
            'error': 'Not Found',
            'message': 'The requested resource was not found.'
        }), 404

    @app.errorhandler(500)
    def handle_500_error(error):
        """Handle 500 Internal Server Error"""
        logger.error(f"500 Internal Server Error: {error}")
        db.session.rollback()  # Rollback any pending transactions
        return jsonify({
            'error': 'Internal Server Error',
            'message': 'An unexpected error occurred. Please try again later.'
        }), 500

    @app.errorhandler(Exception)
    def handle_unexpected_error(error):
        """Handle unexpected errors to prevent worker crashes"""
        logger.error(f"Unexpected error: {error}")
        try:
            db.session.rollback()  # Rollback any pending transactions
        except:
            pass
        return jsonify({
            'error': 'Unexpected Error',
            'message': 'An unexpected error occurred.'
        }), 500

    # Socket.IO error handler
    @socketio.on_error
    def handle_socketio_error(e):
        """Handle Socket.IO errors"""
        logger.error(f"Socket.IO error: {e}")
        try:
            db.session.rollback()
        except:
            pass

    # Database teardown for request context
    @app.teardown_request
    def teardown_request(exception):
        """Clean up database session after each request"""
        if exception:
            logger.error(f"Request exception: {exception}")
            try:
                db.session.rollback()
            except:
                pass
        try:
            db.session.remove()
        except:
            pass

    # Logging setup removed for Vercel deployment

    return app

# Create the Flask app instance for gunicorn
app = create_app()

# Vercel serverless function entry point
# The app instance is created at module level for Vercel
# Vercel handles the server execution automatically

